"""importar_retenciones_as.py - Procesa Excel de retenciones para importar a Account Soft."""

import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import xlrd
from datetime import datetime
import csv


def seleccionar_archivo_excel():
    """Abre un diálogo para seleccionar archivo Excel."""
    root = tk.Tk()
    root.withdraw()  # Ocultar ventana principal
    root.attributes('-topmost', True)  # Traer al frente

    archivo = filedialog.askopenfilename(
        title="Seleccione el archivo Excel de retenciones",
        filetypes=[
            ("Archivos Excel", "*.xlsx *.xls"),
            ("Todos los archivos", "*.*")
        ]
    )

    root.destroy()
    return archivo


def validar_fecha(fecha_str, es_xls=False):
    """Valida y convierte fecha a formato DD/MM/YYYY."""
    if not fecha_str:
        return None

    # Si es un objeto datetime de Excel
    if isinstance(fecha_str, datetime):
        return fecha_str.strftime("%d/%m/%Y")

    # Si es número de serie de fecha de .xls
    if es_xls and isinstance(fecha_str, (int, float)):
        try:
            fecha = xlrd.xldate_as_datetime(fecha_str, 0)
            return fecha.strftime("%d/%m/%Y")
        except:
            pass

    # Si es string, intentar parsear
    try:
        # Intentar varios formatos comunes
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                fecha = datetime.strptime(str(fecha_str), fmt)
                return fecha.strftime("%d/%m/%Y")
            except ValueError:
                continue
        return None
    except:
        return None


def validar_importe(importe_str):
    """Valida y formatea importe."""
    if not importe_str:
        return None

    try:
        # Convertir a float y formatear con 2 decimales
        importe = float(str(importe_str).replace(',', '.'))
        return f"{importe:.2f}"
    except:
        return None


def leer_excel_xls(ruta_archivo):
    """Lee archivos .xls (formato antiguo) usando xlrd."""
    try:
        print(f"\n[...] Cargando archivo .xls: {os.path.basename(ruta_archivo)}")
        wb = xlrd.open_workbook(ruta_archivo, formatting_info=False)
        ws = wb.sheet_by_index(0)

        retenciones = []
        errores = []

        # Leer desde fila 2 (índice 1, asumiendo fila 0 es encabezado)
        for idx in range(1, ws.nrows):
            row = ws.row_values(idx)

            # Columnas: G=6, H=7, I=8, J=9 (índice 0-based)
            fecha = row[6] if len(row) > 6 else None
            numero_cert = row[7] if len(row) > 7 else None
            descripcion = row[8] if len(row) > 8 else None
            importe = row[9] if len(row) > 9 else None

            # Validar que tenga datos
            if not any([fecha, numero_cert, descripcion, importe]):
                continue  # Fila vacía

            # Validar datos (pasando flag es_xls=True para fechas)
            fecha_valid = validar_fecha(fecha, es_xls=True)
            importe_valid = validar_importe(importe)

            if not fecha_valid:
                errores.append(f"Fila {idx + 1}: Fecha inválida '{fecha}'")
                continue

            if not numero_cert:
                errores.append(f"Fila {idx + 1}: Número de certificado vacío")
                continue

            if not importe_valid:
                errores.append(f"Fila {idx + 1}: Importe inválido '{importe}'")
                continue

            retenciones.append({
                'numero_certificado': str(numero_cert).strip(),
                'fecha': fecha_valid,
                'descripcion': str(descripcion).strip() if descripcion else '',
                'importe': importe_valid
            })

        # Mostrar errores si los hay
        if errores:
            print(f"\n[!] Se encontraron {len(errores)} errores:")
            for error in errores[:5]:  # Mostrar máximo 5
                print(f"    {error}")
            if len(errores) > 5:
                print(f"    ... y {len(errores) - 5} más")

        print(f"[OK] {len(retenciones)} retenciones válidas encontradas")
        return retenciones

    except Exception as e:
        print(f"[X] Error al leer el archivo .xls: {e}")
        return None


def leer_excel_xlsx(ruta_archivo):
    """Lee archivos .xlsx (formato nuevo) usando openpyxl."""
    try:
        print(f"\n[...] Cargando archivo .xlsx: {os.path.basename(ruta_archivo)}")
        wb = load_workbook(ruta_archivo, data_only=True)
        ws = wb.active

        retenciones = []
        errores = []

        # Leer desde fila 2 (asumiendo fila 1 es encabezado)
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Columnas: G=6, H=7, I=8, J=9 (índice 0-based)
            fecha = row[6] if len(row) > 6 else None
            numero_cert = row[7] if len(row) > 7 else None
            descripcion = row[8] if len(row) > 8 else None
            importe = row[9] if len(row) > 9 else None

            # Validar que tenga datos
            if not any([fecha, numero_cert, descripcion, importe]):
                continue  # Fila vacía

            # Validar datos
            fecha_valid = validar_fecha(fecha, es_xls=False)
            importe_valid = validar_importe(importe)

            if not fecha_valid:
                errores.append(f"Fila {idx}: Fecha inválida '{fecha}'")
                continue

            if not numero_cert:
                errores.append(f"Fila {idx}: Número de certificado vacío")
                continue

            if not importe_valid:
                errores.append(f"Fila {idx}: Importe inválido '{importe}'")
                continue

            retenciones.append({
                'numero_certificado': str(numero_cert).strip(),
                'fecha': fecha_valid,
                'descripcion': str(descripcion).strip() if descripcion else '',
                'importe': importe_valid
            })

        wb.close()

        # Mostrar errores si los hay
        if errores:
            print(f"\n[!] Se encontraron {len(errores)} errores:")
            for error in errores[:5]:  # Mostrar máximo 5
                print(f"    {error}")
            if len(errores) > 5:
                print(f"    ... y {len(errores) - 5} más")

        print(f"[OK] {len(retenciones)} retenciones válidas encontradas")
        return retenciones

    except Exception as e:
        print(f"[X] Error al leer el archivo .xlsx: {e}")
        return None


def leer_excel_retenciones(ruta_archivo):
    """
    Lee el Excel de retenciones y extrae las columnas necesarias.
    Detecta automáticamente el formato (.xls o .xlsx).

    Columnas:
    - H: Número Certificado
    - G: Fecha Ret./Perc.
    - I: Descripción Operación
    - J: Importe Ret./Perc.

    Returns:
        list: Lista de diccionarios con los datos procesados
    """
    # Detectar formato por extensión
    extension = os.path.splitext(ruta_archivo)[1].lower()

    if extension == '.xls':
        return leer_excel_xls(ruta_archivo)
    elif extension == '.xlsx':
        return leer_excel_xlsx(ruta_archivo)
    else:
        print(f"[X] Formato de archivo no soportado: {extension}")
        print("    Use archivos .xls o .xlsx")
        return None


def guardar_csv_retenciones(retenciones, ruta_salida):
    """Guarda las retenciones en formato CSV."""
    try:
        with open(ruta_salida, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(
                f,
                fieldnames=['numero_certificado', 'fecha', 'descripcion', 'importe'],
                delimiter=';'
            )
            writer.writeheader()
            writer.writerows(retenciones)

        print(f"[OK] Archivo CSV generado: {ruta_salida}")
        return True
    except Exception as e:
        print(f"[X] Error al guardar CSV: {e}")
        return False


def procesar_retenciones(nombre_usuario):
    """
    Proceso principal: selecciona Excel, procesa datos y genera CSV.

    Args:
        nombre_usuario: Nombre del usuario para organizar archivos
    """
    print("\n=== Importar Retenciones a Account Soft ===")
    print("[...] Abriendo selector de archivos...")

    # Seleccionar archivo
    ruta_excel = seleccionar_archivo_excel()

    if not ruta_excel:
        print("[X] No se seleccionó ningún archivo.")
        return False

    if not os.path.exists(ruta_excel):
        print(f"[X] El archivo no existe: {ruta_excel}")
        return False

    # Leer y procesar Excel
    retenciones = leer_excel_retenciones(ruta_excel)

    if not retenciones:
        print("[X] No se pudieron procesar retenciones.")
        return False

    if len(retenciones) == 0:
        print("[X] No se encontraron retenciones válidas en el archivo.")
        return False

    # Crear directorio de salida
    ruta_base = "D:/Users/Lazzarini1/Desktop/D/ex_mis_docu-D/a_PLANILLAS"
    ruta_salida_dir = os.path.join(ruta_base, nombre_usuario, "Retenciones_AS")
    os.makedirs(ruta_salida_dir, exist_ok=True)

    # Generar nombre de archivo con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo_salida = os.path.join(
        ruta_salida_dir,
        f"retenciones_as_{timestamp}.csv"
    )

    # Guardar CSV
    if guardar_csv_retenciones(retenciones, archivo_salida):
        print("\n" + "="*60)
        print("✓ PROCESO COMPLETADO")
        print("="*60)
        print(f"Total de retenciones: {len(retenciones)}")
        print(f"Archivo generado: {archivo_salida}")
        print("\nEl archivo está listo para importar a Account Soft")
        print("(La automatización de importación se agregará en una próxima versión)")
        print("="*60)
        return True

    return False
